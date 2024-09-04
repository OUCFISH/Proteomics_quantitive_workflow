import re
import os
import paddleocr
from paddleocr import PaddleOCR
import time
from openpyxl import load_workbook
import pandas as pd
import shutil
import argparse


def create_pro_info():
    if not os.path.exists('./table'):
        os.mkdir('./table')
    if not os.path.exists('./image'):
        os.mkdir('./image')
    def ocr(pdf_file):
        texts = []
        PAGE_NUM = 1
        ocr = PaddleOCR(use_angle_cls=True, lang="ch", page_num=PAGE_NUM)
        result = ocr.ocr(pdf_file, cls=True)
        for idx in range(len(result)):
            res = result[idx]
            txts = [line[1][0] for line in res]
        texts += txts

        all_text = ' '.join(texts)

        info_dict = {}
        info_dict['项目编号'] = os.path.basename(pdf_file).replace('.pdf', '')
        info_dict['合同号'] = re.findall('合同编号[：:;]{1,2}\s*(\S*)', all_text)[0] if re.findall('合同编号[：:;]{1,2}\s*(\S*)',
                                                                                           all_text) else ''
        info_dict['项目名称'] = 'Library-free DIA蛋白质组学分析'
        info_dict['项目类别'] = '医学蛋白质组学分析'
        info_dict['项目类型'] = '标准分析'
        info_dict['完成日期'] = time.strftime("%Y/%m/%d", time.localtime())
        info_dict['客户姓名'] = re.findall('客户姓名[：:;]{1,2}\s*(\S*)', all_text)[0] if re.findall('客户姓名[：:;]{1,2}\s*(\S*)',
                                                                                            all_text) else ''
        info_dict['委托单位'] = re.findall('客户单位[：:;]{1,2}\s*(\S*)', all_text)[0] if re.findall('客户单位[：:;]{1,2}\s*(\S*)',
                                                                                            all_text) else ''
        return info_dict

    pdf_file = [fi for fi in os.listdir(r'./result/1_Sampleinfo') if fi.endswith('.pdf')][0]
    infos = ocr(os.path.join(r'./result/1_Sampleinfo', pdf_file))
    with open('./table/project_info.txt', 'w', encoding='utf-8') as f:
        for k in ['项目编号', '合同号', '项目名称', '项目类别', '项目类型', '完成日期', '客户姓名', '委托单位']:
            f.write(f'{k}\t{infos[k]}\n')


def create_sample_info():
    sample_xlsx = [fi for fi in os.listdir(r'./result/1_Sampleinfo') if
                   fi.endswith('.xlsx') and ('GroupSheet' in fi or '样品分组' in fi) and '$' not in fi][0]
    sample_info = load_workbook(os.path.join(r'./result/1_Sampleinfo', sample_xlsx))
    sheet = sample_info['样品分组表']
    if not sheet['B4'].value:
        sheet = sample_info['分组表示例']
    Species = sheet['E5'].value
    i = 12
    samples = []
    while True:
        if sheet[f'A{i}'].value:
            if sheet[f'C{i}'].value:
                group = sheet[f'C{i}'].value
            samples.append([sheet[f'A{i}'].value, group, Species])
            i += 1
        else:
            break
    with open('./table/sample_info.txt', 'w', encoding='utf-8') as f:
        f.write('Sample Names\tGroup Names\tSpecies')
        for s in samples:
            f.write(f'\n{s[0]}\t{s[1]}\t{s[2]}')


def create_search_result():
    Peptides_file = \
    [fi for fi in os.listdir(r'./result/2_Data_Summary') if fi.endswith('.xlsx') and 'Peptide' in fi and '$' not in fi][
        0]
    df = pd.read_excel(os.path.join(r'./result/2_Data_Summary', Peptides_file))
    Peptides = len(df)
    Identified_file = [fi for fi in os.listdir(r'./result/2_Data_Summary') if
                       fi.endswith('.xlsx') and 'ProteinGroup' in fi and '$' not in fi][0]
    df = pd.read_excel(os.path.join(r'./result/2_Data_Summary', Identified_file))
    Identified = len(df)
    Quantified_file = [fi for fi in os.listdir(r'./result/2_Data_Summary') if
                       fi.endswith('.xlsx') and 'SummaryStat' in fi and '$' not in fi][0]
    df = pd.read_excel(os.path.join(r'./result/2_Data_Summary', Quantified_file))
    Quantified = len(df)
    with open('./table/search_result.txt', 'w', encoding='utf-8') as f:
        f.write('Result\tPeptides\tIdentified proteins\tQuantified proteins\n')
        f.write(f'Uniprot database\t{Peptides}\t{Identified}\t{Quantified}')


def create_diff_analyse_result():   # 包括变量参数生成
    # 获取FC及Pvalue
    with open(r'./config/default.conf', 'r', encoding='utf-8') as conf:
        for line in conf:
            if line.startswith('Fold_Change1'):
                FC = line.split(' ')[-1].strip()
            if line.startswith('pvalue'):
                pvalue = line.split(' ')[-1].strip()
                break
    # 获取fasta文件名
    fasta = ' '
    if os.path.exists(r'./data'):
        for file in os.listdir(r'./data'):
            if file.endswith('.fasta'):
                fasta = f'本次数据库为：{file}'
                break

    sample_xlsx = [fi for fi in os.listdir(r'./result/1_Sampleinfo') if
                   fi.endswith('.xlsx') and ('GroupSheet' in fi or '样品分组' in fi) and '$' not in fi][0]
    sample_info = load_workbook(os.path.join(r'./result/1_Sampleinfo', sample_xlsx))
    sheet = sample_info['样品分组表']
    if not sheet['B4'].value:
        sheet = sample_info['分组表示例']
    compares = []
    i = 12
    while True:
        if sheet[f'E{i}'].value:
            compares.append(f'{sheet[f"E{i}"].value}_vs_{sheet[f"F{i}"].value}')
            i += 1
        else:
            break
    with open('variable.txt', 'w', encoding='utf-8', newline='') as f:
        f.write('compare\t')
        f.write('\t'.join(compares))
        f.write('\n')
        f.write('gsea_go')
        for compare in compares:
            f.write(
                f'\t{compare}_GO_GSEA_enrichment_{compare.split("_vs_")[0]}\t{compare}_GO_GSEA_enrichment_{compare.split("_vs_")[1]}')
        f.write('\n')
        f.write('gsea_kegg')
        for compare in compares:
            f.write(
                f'\t{compare}_KEGG_GSEA_enrichment_{compare.split("_vs_")[0]}\t{compare}_KEGG_GSEA_enrichment_{compare.split("_vs_")[1]}')
        f.write('\n')
        f.write('gsea_reactome')
        for compare in compares:
            f.write(
                f'\t{compare}_Reactome_GSEA_enrichment_{compare.split("_vs_")[0]}\t{compare}_Reactome_GSEA_enrichment_{compare.split("_vs_")[1]}')
        f.write('\n')
        f.write('gsea_domain')
        for compare in compares:
            f.write(
                f'\t{compare}_Domain_GSEA_enrichment_{compare.split("_vs_")[0]}\t{compare}_Domain_GSEA_enrichment_{compare.split("_vs_")[1]}')
        f.write('\n')
        f.write('cate_go\tBP\tCC\tMF\n')
        f.write(f'FC\t{FC}\n')
        f.write(f'pvalue\t{pvalue}\n')
        f.write(f'fasta\t{fasta}\n')
        f.write(f'AA\t{"AA" if args.advanced_analyse else "noAA"}\n')
    xlsx = [fi for fi in os.listdir(r'./result/2_Data_Summary') if
            fi.endswith('.xlsx') and 'SummaryStat' in fi and '$' not in fi]
    df = pd.read_excel(os.path.join(r'./result/2_Data_Summary', xlsx[0]))
    with open(r'./table/diff_analyse_result.txt', 'w', encoding='utf-8') as f:
        f.write(f'Statistical methods\tFold change and student’s t test\tFold change > {FC}\tP value < {pvalue}\n')
        f.write('Comparisons\tUp Regulation\tDown Regulation\tAll\n')
        for compare in compares:
            up = df[df[f"{compare}.Regulation"] == "Up"].shape[0]
            down = df[df[f"{compare}.Regulation"] == "Down"].shape[0]
            f.write(
                f'{compare}\t{up}\t{down}\t{up + down}\n')
    with open(r'./table/diff_exp.txt', 'w', encoding='utf-8', newline='') as f:
        f.write('Control_vs_Treat\tUp Regulation\tDown Regulation\tTotal\n')
        for compare in compares:
            up = df[df[f"{compare}.Regulation"] == "Up"].shape[0]
            down = df[df[f"{compare}.Regulation"] == "Down"].shape[0]
            f.write(f'{compare}\t{up}\t{down}\t{up + down}\n')


def create_annotations():
    xlsx = [fi for fi in os.listdir(r'./result/2_Data_Summary') if
            fi.endswith('.xlsx') and 'SummaryStat' in fi and '$' not in fi][0]
    df = pd.read_excel(os.path.join(r'./result/2_Data_Summary', xlsx))
    total = len(df)
    GO = df[df['Gene.Ontology.(GO)'].notnull()].shape[0]
    KEGG = df[df['KEGG'].notnull()].shape[0]
    # Reactome = df[df['Reactome'].notnull()].shape[0]
    Interproscan = df[df['InterPro'].notnull()].shape[0]
    TF = df[df['TF'].notnull()].shape[0]
    Kinase = df[df['kinase']].shape[0]
    SL = df[df['subcellular_location'].notnull()].shape[0]
    eggNOG = df[df['eggNOG'].notnull()].shape[0]
    DO = df[df['DO'].notnull()].shape[0]
    Wikipathway = df[df['Wikipathway'].notnull()].shape[0]
    with open(r'./table/annotations.txt', 'w', encoding='utf-8', newline='') as f:
        f.write('Database\tAnnotated\tPercent\n')
        f.write(f'GO\t{GO}\t{round(GO / total * 100, 2):.2f}%\n')
        f.write(f'KEGG\t{KEGG}\t{round(KEGG / total * 100, 2):.2f}%\n')
        # f.write(f'Reactome\t{Reactome}\t{round(Reactome / total * 100, 2):.2f}%\n')
        f.write(f'Domain\t{Interproscan}\t{round(Interproscan / total * 100, 2):.2f}%\n')
        f.write(f'TF\t{TF}\t{round(TF / total * 100, 2):.2f}%\n')
        f.write(f'Kinase\t{Kinase}\t{round(Kinase / total * 100, 2):.2f}%\n')
        f.write(f'Subcellular Localization\t{SL}\t{round(SL / total * 100, 2):.2f}%\n')
        f.write(f'eggNOG\t{eggNOG}\t{round(eggNOG / total * 100, 2):.2f}%\n')
        f.write(f'DO\t{DO}\t{round(DO / total * 100, 2):.2f}%\n')
        f.write(f'Wikipathway\t{Wikipathway}\t{round(Wikipathway / total * 100, 2):.2f}%\n')


def style():
    def header_format(info):
        if info == 'base_info':
            bg_color = '#0586C0'
        elif info == 'sample_info':
            bg_color = '#F5780C'
        elif info == 'Regulation':
            bg_color = '#8064A2'
        elif info == 'annotation':
            bg_color = '#595959'
        else:
            bg_color = '#0586C0'
        return {
            'font_name': 'calibri',
            'bold': False,  # 设置字体加粗
            'font_size': 11,  # 设置字体大小
            'bg_color': bg_color,  # 设置背景颜色
            'color': 'white',  # 设置字体颜色
            'align': 'left',  # 设置文本居中对齐
            'valign': 'vcenter',  # 设置垂直居中对齐
            'border_color': '#CCCCCC',  # 设置边框颜色
            'border': 1,  # 边框宽度
            'top': 1,  # 上边框
            'left': 1,  # 左边框
            'right': 1,  # 右边框
            'bottom': 1  # 底边框
        }

    def cell_format(info):
        if info == 'Down':
            color = '#0070C0'
        elif info == 'Up':
            color = '#FF0000'
        else:
            color = 'black'
        return {
            'font_name': 'calibri',
            'bold': False,  # 设置字体加粗
            'font_size': 10,  # 设置字体大小
            'bg_color': '#FFFFFF',  # 设置背景颜色
            'color': color,  # 设置字体颜色
            'border_color': '#CCCCCC',  # 设置边框颜色
            'border': 1,  # 边框宽度
            'top': 1,  # 上边框
            'left': 1,  # 左边框
            'right': 1,  # 右边框
            'bottom': 1  # 底边框
        }

    xlsx = [fi for fi in os.listdir(r'./result/2_Data_Summary') if
            fi.endswith('.xlsx') and 'SummaryStat' in fi and '$' not in fi][0]
    xlsx_path = rf'./result/2_Data_Summary/{xlsx}'
    df = pd.read_excel(xlsx_path)
    writer = pd.ExcelWriter(xlsx_path.replace('.xlsx', '_styled.xlsx'), engine='xlsxwriter')
    workbook = writer.book
    ws = workbook.add_worksheet('Sheet 1')
    columns = df.columns.tolist()
    info = 'base_info'
    for i in range(len(columns)):
        if i > 0 and (columns[i - 1] == 'Coverage' or columns[i - 1].endswith('.Regulation')):
            info = 'sample_info'
        if columns[i].endswith('.Regulation'):
            info = 'Regulation'
        if columns[i] == ('Gene.Ontology.(GO)'):
            info = 'annotation'
        ws.write(0, i, columns[i], workbook.add_format(header_format(info)))
    ws.set_row(0, height=20)
    df = df.fillna('')
    rows, cols = df.shape
    for row in range(rows):
        for col in range(cols):
            value = df.iat[row, col]
            if col == cols - 2:
                ws.write(row + 1, col, str(value), workbook.add_format(cell_format(value)))
            else:
                ws.write(row + 1, col, value, workbook.add_format(cell_format(value)))
    writer._save()


def get_6_enrichments():
    path = r'./result/6_GSEA'
    folder = os.listdir(path)
    for dir in folder:
        compares = os.listdir(os.path.join(path, dir))
        for compare in compares:
            if not os.path.isdir(os.path.join(path, dir, compare)):
                continue
            files = [fi for fi in os.listdir(os.path.join(path, dir, compare)) if 'enrichment' in fi.lower()]
            for file in files:
                shutil.copy2(os.path.join(path, dir, compare, file), os.path.join(path, dir, file))



def get_GO_enrichment_table():
    folder = os.listdir(r'./result/5_Enrichment/5.1_GO')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.1_GO/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/go_enrichment.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/go_enrichment.tsv'):
        print('No GO enrichment result')
        with open(r'./table/go_enrichment.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No GO enrichment result')


def get_KEGG_enrichment_table():
    folder = os.listdir(r'./result/5_Enrichment/5.2_KEGG')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.2_KEGG/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/kegg_enrichment.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/kegg_enrichment.tsv'):
        print('No KEGG enrichment result')
        with open(r'./table/kegg_enrichment.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No KEGG enrichment result')


def get_TF_table():
    folder = os.listdir(r'./result/7_TF')
    df = pd.read_excel(r'./result/7_TF/tf.family.xlsx')
    df.to_csv(r'./table/TF_total.tsv', sep='\t', index=False)
    for i in range(len(folder)):
        try:
            path = rf'./result/7_TF/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/TF.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/TF.tsv'):
        print('No TF result')
        with open(r'./table/TF.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No TF result')


def get_subcellular_location_table():
    folder = os.listdir(r'./result/8_Subcellular-location')
    for i in range(len(folder)):
        try:
            path = rf'./result/8_Subcellular-location/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('sub.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/sub.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/sub.tsv'):
        print('No subcellular location result')
        with open(r'./table/sub.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No subcellular location result')


def get_PPI_table():
    folder = os.listdir(r'./result/9_PPI')
    for i in range(len(folder)):
        try:
            path = rf'./result/9_PPI/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('network.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/ppi.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/ppi.tsv'):
        print('No PPI result')
        with open(r'./table/ppi.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No PPI result')


def get_kinase_table():
    folder = os.listdir(r'./result/10_Kinase')
    for i in range(len(folder)):
        try:
            path = rf'./result/10_Kinase/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('kinase.gene.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/kinase.tsv', sep='\t', index=False)
            return
        except:
            continue
    with open(r'./table/kinase.tsv', 'w') as f:
        f.write('The differential experssion protein not be annotated in Kinase database')


def get_reactome_enrichment_table():
    folder = os.listdir(r'./result/5_Enrichment/5.4_Reactome')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.4_Reactome/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/Reactome_enrichment.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/Reactome_enrichment.tsv'):
        print('No Reactome enrichment result')
        with open(r'./table/Reactome_enrichment.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No Reactome enrichment result')


def get_domain_table():
    folder = os.listdir(r'./result/5_Enrichment/5.3_domain')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.3_domain/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/domain.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/domain.tsv'):
        print('No domain enrichment result')
        with open(r'./table/domain.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No domain enrichment result')


def get_DO_table():
    folder = os.listdir(r'./result/5_Enrichment/5.5_DO')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.5_DO/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/DO_enrichment.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/DO_enrichment.tsv'):
        print('No DO enrichment result')
        with open(r'./table/DO_enrichment.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No DO enrichment result')


def get_wikipathway_table():
    folder = os.listdir(r'./result/5_Enrichment/5.6_WikiPathway')
    for i in range(len(folder)):
        try:
            path = rf'./result/5_Enrichment/5.6_WikiPathway/{folder[i]}'
            xls = [fi for fi in os.listdir(path) if fi.endswith('enrichment.xlsx')]
            df = pd.read_excel(os.path.join(path, xls[0]))
            df.to_csv(r'./table/wikipathway_enrichment.tsv', sep='\t', index=False)
            break
        except:
            continue
    if not os.path.exists(r'./table/wikipathway_enrichment.tsv'):
        print('No wikipathway enrichment result')
        with open(r'./table/wikipathway_enrichment.tsv', 'w', encoding='utf-8', newline='') as f:
            f.write('No wikipathway enrichment result')


def get_KEGG_pathway_pic():
    compares = os.listdir(r'./result/5_Enrichment/5.2_KEGG')
    for compare in compares:
        path = rf'./result/5_Enrichment/5.2_KEGG/{compare}/{compare}_KEGG_Pathway_plot/{compare}_KEGG_Pathway_plot'
        for fi in os.listdir(path):
            if fi.endswith('.png'):
                shutil.copy2(os.path.join(path, fi), rf'./image/{compare}_KEGG_pathway.png')
                break
        filename1 = f'{compare}_GO_GSEA_enrichment_{compare.split("_vs_")[0]}.xls'
        filename2 = f'{compare}_GO_GSEA_enrichment_{compare.split("_vs_")[1]}.xls'
        try:
            shutil.copy2(rf'./result/6_GSEA/6.1_GO_GSEA/{compare}/{filename1}',
                         rf'./result/6_GSEA/6.1_GO_GSEA/{filename1}')
            shutil.copy2(rf'./result/6_GSEA/6.1_GO_GSEA/{compare}/{filename2}',
                         rf'./result/6_GSEA/6.1_GO_GSEA/{filename2}')
        except:
            pass
        filename1 = f'{compare}_KEGG_GSEA_enrichment_{compare.split("_vs_")[0]}.xls'
        filename2 = f'{compare}_KEGG_GSEA_enrichment_{compare.split("_vs_")[1]}.xls'
        try:
            shutil.copy2(rf'./result/6_GSEA/6.2_KEGG_GSEA/{compare}/{filename1}',
                         rf'./result/6_GSEA/6.2_KEGG_GSEA/{filename1}')
            shutil.copy2(rf'./result/6_GSEA/6.2_KEGG_GSEA/{compare}/{filename2}',
                         rf'./result/6_GSEA/6.2_KEGG_GSEA/{filename2}')
        except:
            pass


def dealing_errpr_infos():
    path = r'./result/4_Diff_Expressed/4.2_Diff_Repeat'
    if os.path.exists(rf'{path}/error_info.png'):
        shutil.copy(rf'{path}/error_info.png', rf'{path}/upset_plot.png')
    if os.path.exists(rf'{path}/error_info.png'):
        shutil.copy(rf'{path}/error_info.png', rf'{path}/venn.png')
    path = r'./result/4_Diff_Expressed/4.3_Volcano'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/kinase-barplot.png')
    path = r'./result/4_Diff_Expressed/4.4_Heatmap'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/heatmap.png')
    path = r'./result/4_Diff_Expressed/4.5_Trend'
    if os.path.exists(rf'{path}/error_info.png'):
        shutil.copy(rf'{path}/error_info.png', rf'{path}/K_Means_Plot.png')
    path = r'./result/4_Diff_Expressed/4.6_Radar'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/Radar.png')
    path = r'./result/5_Enrichment/5.1_GO'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/GO-barplot.jpg')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/GO_dag.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GO_bar.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GO_circos.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GO_network.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GO_richfactor.png')
        for cate in ['BP', 'CC', 'MF']:
            if os.path.exists(rf'{path}/{folder[i]}/{cate}/error_info.png'):
                shutil.copy(rf'{path}/{folder[i]}/{cate}/error_info.png',
                            rf'{path}/{folder[i]}/{cate}/GO-chrodplot.jpg')
                shutil.copy(rf'{path}/{folder[i]}/{cate}/error_info.png',
                            rf'{path}/{folder[i]}/{cate}/GO-updown-barplot.jpg')
    path = r'./result/5_Enrichment/5.2_KEGG'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_KEGG_bar.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_KEGG_circos.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_KEGG_network.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_KEGG_richfactor.png')
    path = r'./result/5_Enrichment/5.3_domain'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/interPro-barplot.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/interPro-bubbleplot.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_interPro_circos.png')
    path = r'./result/5_Enrichment/5.4_Reactome'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_Reactome_bar.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_Reactome_circos.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_Reactome_network.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png',
                        rf'{path}/{folder[i]}/{folder[i]}_Reactome_richfactor.png')
    path = r'./result/6_GSEA/6.1_GO_GSEA'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GO_Global_ES.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png',
                        rf'{path}/{folder[i]}/{folder[i]}_GO_p_values_vs_NES.png')
    path = r'./result/6_GSEA/6.2_KEGG_GSEA'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_KEGG_Global_ES.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png',
                        rf'{path}/{folder[i]}/{folder[i]}_KEGG_p_values_vs_NES.png')
    # path = r'./result/7_TF'
    # folder = os.listdir(path)
    # for i in range(len(folder)):
    #     if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
    #         shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}\tf-analysis.png')
    path = r'./result/8_Subcellular-location'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/sub-barplot.png')
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/sub-pieplot.png')
    path = r'./result/9_PPI'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}.network.png')
    path = r'./result/10_Kinase'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/kinase-barplot.png')
    path = r'./result/11_AdcancedPlot/11.4_GSEA-multiPathwayPlot/GO'
    folder = os.listdir(path)
    for i in range(len(folder)):
        if os.path.exists(rf'{path}/{folder[i]}/error_info.png'):
            shutil.copy(rf'{path}/{folder[i]}/error_info.png', rf'{path}/{folder[i]}/{folder[i]}_GSEA.png')


def xlsx_2_tsv():
    file_path = r'./result/3_Quality_Control/3.2_Quantitative_Statistics/3.2.8_PCA/PCA.xlsx'
    df = pd.read_excel(file_path)
    df.to_csv(file_path.replace('.xlsx','.tsv'), sep='\t', index=False)
    file_path = r'./result/4_Diff_Expressed/4.5_Trend/kmeans.xlsx'
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        df.to_csv(file_path.replace('.xlsx', '.tsv'), sep='\t', index=False)
    else:
        with open(file_path.replace('.xlsx', '.tsv'), 'w', encoding='utf-8') as f:
            f.write('       Sample grouping is less than 3, unable to conduct analysis      ')

    file_path = r'./result/3_Quality_Control/3.2_Quantitative_Statistics/3.2.6_Correlation_Analysis/correlation.xlsx'
    df = pd.read_excel(file_path)
    df.to_csv(file_path.replace('.xlsx','.tsv'), sep='\t', index=False)

#抓取6_GSEA/p_top20/里面的第一张png
def get_GSEA_pic():
    list_dir = ['6.1_GO_GSEA','6.2_KEGG_GSEA','6.3_Reactome_GSEA','6.4_Domain_GSEA']
    for dir in list_dir:
        compares = os.listdir(rf'./result/6_GSEA/{dir}')
        for compare in compares:
            if not os.path.isdir(rf'./result/6_GSEA/{dir}/{compare}'):
                continue
            path = rf'./result/6_GSEA/{dir}/{compare}/p_top20/'
            for fi in os.listdir(path):
                if fi.endswith('.png'):
                    shutil.copy2(os.path.join(path, fi), rf'./image/{dir}_{compare}.png')
                    break


def create_md():
    with open('prot_report_DIA.md', 'w', encoding='utf-8', newline='') as md:
        with open(r'./temp/prot_report_test_DIA.md', 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip() == 'Advanced_analytics':
                    if args.advanced_analyse:
                        with open(r'./temp/Advanced_analytics.md', 'r', encoding='utf-8') as fi:
                            for l in fi:
                                md.write(l)
                else:
                    md.write(line)


def main():
    create_md()
    create_pro_info()
    create_sample_info()
    create_search_result()
    create_diff_analyse_result()
    create_annotations()
    get_GO_enrichment_table()
    get_KEGG_enrichment_table()
    get_KEGG_pathway_pic()
    get_reactome_enrichment_table()
    get_TF_table()
    get_subcellular_location_table()
    get_PPI_table()
    get_kinase_table()
    get_domain_table()
    get_DO_table()
    get_wikipathway_table()
    get_6_enrichments()
    dealing_errpr_infos()
    get_GSEA_pic()
    xlsx_2_tsv()
    style()



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Pre report")
    parser.add_argument('--advanced-analyse', help='enable advanced analyse', action='store_true')
    args = parser.parse_args()
    main()